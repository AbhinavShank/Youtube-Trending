import pandas as pd  # 处理数据
import matplotlib.pyplot as plt  # 画图
from matplotlib import cm
import json
import nltk
import seaborn as sns
import datetime
from nltk.tokenize import sent_tokenize
from nltk.tokenize import word_tokenize
from wordcloud import WordCloud, STOPWORDS
from openpyxl import load_workbook

# nltk.download()
'''
1. Video ID
2. Trending date
3. Title
4. Channel title
5. Category ID
6. Publish time
7. Tags
8. Views
9. Likes
10. Dislikes
11. Comments count
12. Description
'''

Category_dict = {'1': 'Film & Animation',
                 '2': 'Autos & Vehicles',
                 '10': 'Music',
                 '15': 'Pets & Animals',
                 '17': 'Sports',
                 '18': 'Short Movies',
                 '19': 'Travel & Events',
                 '20': 'Gaming',
                 '21': 'Videoblogging',
                 '22': 'People & Blogs',
                 '23': 'Comedy',
                 '24': 'Entertainment',
                 '25': 'News & Politics',
                 '26': 'Howto & Style',
                 '27': 'Education',
                 '28': 'Science & Technology',
                 '29': 'Nonprofits & Activism',
                 '30': 'Movies',
                 '31': 'Anime/Animation',
                 '32': 'Action/Adventure',
                 '33': 'Classics',
                 '34': 'Comedy',
                 '35': 'Documentary',
                 '36': 'Drama',
                 '37': 'Family',
                 '38': 'Foreign',
                 '39': 'Horror',
                 '40': 'Sci-Fi/Fantasy',
                 '41': 'Thriller',
                 '42': 'Shorts',
                 '43': 'Shows',
                 '44': 'Trailers'}

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
result_df = pd.DataFrame()
VideoID_list = []
TrendingDate_list = []
Title_list = []
Channeltitle_list = []
CategoryID_list = []
Publishtime_list = []
Tags_list = []
Views_list = []
Likes_list = []
Dislikes_list = []
Commentscount_list = []
Category_list = []
Description_list = []
Tags_list = []
with open("USvideos_New.json", 'r') as load_f:
    load_dict = json.load(load_f)

for each_dict in load_dict:
    for name, value in each_dict.items():
        if name == 'video_id':
            VideoID_list.append(value)
        elif name == 'trending_date':
            TrendingDate_list.append(value)
        elif name == 'title':
            Title_list.append(value)
        elif name == 'channel_tittle':
            Channeltitle_list.append(value)
        elif name == 'category_id':
            CategoryID_list.append(value)
            Category_list.append(Category_dict[value])
        elif name == 'publish_time':
            Publishtime_list.append(value)
        elif name == 'views':
            Views_list.append(value)
        elif name == 'likes':
            Likes_list.append(value)
        elif name == 'dislikes':
            Dislikes_list.append(value)
        elif name == 'comment_count':
            Commentscount_list.append(value)
        elif name == 'description':
            Description_list.append(value)
        elif name == 'tags':
            Tags_list.append(value)

result_df['video_id'] = VideoID_list
result_df['trending_date'] = TrendingDate_list
result_df['title'] = Title_list
result_df['channel_tittle'] = Channeltitle_list
result_df['category_id'] = CategoryID_list
result_df['publish_time'] = Publishtime_list
result_df['views'] = Views_list
result_df['likes'] = Likes_list
result_df['dislikes'] = Dislikes_list
result_df['comment_count'] = Commentscount_list
result_df['Category_list'] = Category_list
result_df['description'] = Description_list
result_df['tags'] = Tags_list

result_df["description"] = result_df["description"].fillna(value="")
writer = pd.ExcelWriter('USUSUSDATA.xlsx')
result_df = result_df.drop(index=[46, 133])  # 6，65，199空数值
result_df.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=0, index=False)  # 一开始生成的数字 不要
writer.save()


def handle_excel():
    import re
    # 处理Excel  预处理文字信息 包含了符号  外文
    wb = load_workbook('USUSUSDATA.xlsx')  # 打开Excel
    ws = wb.active  # 得到活动的work sheet
    # 把一些非ASCII字符，替换成对应的ASCII
    change_dict = {'–': '-',
                   '—': '-',
                   '’': "'",
                   '‘': "'",
                   '“': '"',
                   '”': '"',
                   '！': '!',
                   '│': '|',
                   '…': '...',
                   '•': '·',
                   '️': ' ',
                   }

    columns = ['C', 'D', 'L']

    strange_characters = set()  # 集合，存储文本中检测到的所有非ASCII字符  没有重复
    # remove_list = ['facebook', 'twitter', 'Images', 'Instagram', 'YouTube', 'http', 'video']

    for column in columns:  # Excel表中的C D K 三列数据进行处理
        for cell in ws[column]:
            # 预处理，把change_dict中一些非ASCII字符，替换成对应的ASCII
            if cell.value:  # 不为空
                for key, value in change_dict.items():
                    cell.value = cell.value.replace(key, value)
                origin = cell.value

                for c in origin:
                    if not c.isascii():
                        strange_characters.add(c)  # 把检测到的所有非ASCII字符存储到strange_characters中

                # ''.join(filter(str.isalnum, origin))   #保留字母数字

                cell.value = ''.join(filter(str.isascii, origin))  # 过滤所有非ASCII字符  x代表参数  x.isascii()
                # if origin != cell.value:  # 打印检测出含有非ASCII字符的Excel中的文本
                # print('原  文：', origin)
                # print('处理后：', cell.value)
                # print()

                # 去除url
                cell.value = re.sub(r'(https|http)?:\/\/(\w|\.|\/|\?|\=|\&|\%)*\b', '', cell.value, flags=re.MULTILINE)

                # 去除facebook twitter之类的东西
                # for key in remove_list:
                #     cell.value = re.sub(key, '', cell.value, flags=re.IGNORECASE)

    # print(strange_characters)  # 打印检测出的所有非ASCII字符

    wb.save('USresults.xlsx')  # 文件另存为results.xlsx, 也可保存为youtube.xlsx，这样可以覆盖原文件


handle_excel()


def fetch_description_column():
    wb = load_workbook('USresults.xlsx')  # 打开Excel
    ws = wb.active  # 得到活动的work sheet
    with open('usdescription.txt', 'w', encoding='utf-8') as f:
        for cell in ws['L']:
            if cell.value:
                f.write(cell.value)
    wb.close()


fetch_description_column()

# mytext_file = open('usdescription.txt', 'w')
# error = 0
# for each in result_df.iloc[:, -2]:  # ：，表示最后一列  只取description那列
#     # print(each)
#     try:
#         mytext_file.write(each)
#     except:
#         print('can\'t read')
#         error += 1
# print('There are %dcan not read' % error)
#
# mytext_file.close()

with open('usdescription.txt', encoding='utf-8') as f:
    words = word_tokenize(f.read())  # words是分词之后的列表  按照文档把每句话 都分开来

# str.lower(),把字符串转为小写
stopwordslist = set(map(str.lower, map(str.strip, open('stop.txt', encoding='UTF-8').readlines())))

# 清理数据
wordslist = []
for word in words:
    if word.lower() not in stopwordslist:  # 保存去掉停用词之后的词
        for w in word:  # 除去含有字符的词，只允许有字母和数字
            if not w.isalnum():
                break
        else:
            wordslist.append(word)

# 统计词频
counts = {}
for word in wordslist:
    if len(word) > 1:
        counts[word] = counts.get(word, 0) + 1
# print(counts)
# mytext_file = open('usdescription.txt', 'r').read()

# background_image = plt.imread('timg.jpg')
top_10 = list(sorted(counts.items(), key=lambda item: item[1], reverse=True))[:10]
# 存入Excel
wb = load_workbook('frequency_of_wrolds.xlsx')
ws = wb.active
for x in range(10):
    ws['E' + str(x + 2)] = top_10[x][0] + '(' + str(top_10[x][1]) + ')'
print(top_10)
wb.save('frequency_of_wrolds.xlsx')




font = r'C:\Windows\Fonts\simfang.ttf'
wordcloud = WordCloud(font_path=font, background_color='black', width=1000, height=800,
                      max_words=200, max_font_size=100)
wordcloud.generate_from_frequencies(counts)
plt.imshow(wordcloud)
plt.axis("off")  # 去掉坐标轴
plt.show()

wordcloud.to_file('USdescription.png')
