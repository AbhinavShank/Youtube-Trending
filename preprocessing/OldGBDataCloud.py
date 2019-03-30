import csv
import re

import matplotlib.pyplot as plt
from nltk.tokenize import word_tokenize
from wordcloud import WordCloud
from openpyxl import load_workbook

# 把一些非ASCII字符，替换成对应的ASCII
change_dict = {'–': '-',
               '—': '-',
               '’': "'",
               '‘': "'",
               '“': '"',
               '”': '"',
               '！': '!',
               '│': '|',
               '…': '...',
               '•': '·',
               '️': ' ',
               }


def word_cloud(file='OLDGBdescription'):
    with open('OLDGBdescription.txt', encoding='utf-8') as f:
        words = word_tokenize(f.read())  # words是分词之后的列表  按照文档把每句话 都分开来

    # str.lower(),把字符串转为小写
    stopwordslist = set(map(str.lower, map(str.strip, open('stop.txt', encoding='UTF-8').readlines())))

    # 清理数据
    wordslist = []
    for word in words:
        if word.lower() not in stopwordslist:  # 保存去掉停用词之后的词
            for w in word:  # 除去含有字符的词，只允许有字母和数字
                if not w.isalnum():
                    break
            else:
                wordslist.append(word)

    # 统计词频
    counts = {}
    for word in wordslist:
        if len(word) > 1:
            counts[word] = counts.get(word, 0) + 1
    # print(counts)
    # mytext_file = open('usdescription.txt', 'r').read()

    # background_image = plt.imread('timg.jpg')
    top_10 = list(sorted(counts.items(), key=lambda item: item[1], reverse=True))[:10]
    # 存入Excel
    wb = load_workbook('词频.xlsx')
    ws = wb.active
    for x in range(10):
        ws['D' + str(x + 2)] = top_10[x][0] + '(' + str(top_10[x][1]) + ')'
    print(top_10)
    wb.save('词频.xlsx')

    font = r'C:\Windows\Fonts\simfang.ttf'
    wordcloud = WordCloud(font_path=font, background_color='black', width=1000, height=800,
                          max_words=200, max_font_size=100)
    wordcloud.generate_from_frequencies(counts)
    plt.imshow(wordcloud)
    plt.axis("off")  # 去掉坐标轴
    plt.show()

    wordcloud.to_file(file + '.png')


def cloud_csv():
    for file in [ 'OldGBvideos']:
        print(file + '.csv')
        with open(file + '.csv', encoding='utf-8') as f:
            fw = open('OLDGBdescription.txt', 'w', encoding='utf-8')
            reader = csv.DictReader(f)
            for date in reader:
                text = date['description']
                # 预处理，把change_dict中一些非ASCII字符，替换成对应的ASCII
                if text:  # 不为空
                    for key, value in change_dict.items():
                        text = text.replace(key, value)

                    # 过滤所有非ASCII字符  x代表参数  x.isascii()
                    text = ''.join(filter(str.isascii, text))
                    # 去除url
                    text = re.sub(r'(https|http)?:\/\/(\w|\.|\/|\?|\=|\&|\%)*\b', '', text, flags=re.MULTILINE).strip()
                    if text:
                        fw.write(text)
            fw.close()
        word_cloud(file)


if __name__ == '__main__':
    cloud_csv()